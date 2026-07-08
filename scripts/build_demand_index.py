"""Build an ONDC category-demand index from REAL transaction data (AIKosh).

Sources:
- ONDC-MyStore_App_Products_Orders.xlsx  (9K products w/ categories, 2K order rows)
- ONDC-WowGeni_App_Orders.xlsx           (137K real order lines)

Products/orders are mapped to ONDC retail domains by keyword; order volumes
aggregate into a High/Medium/Low demand index per domain. Output ships with
the API (apps/api/data/category_demand.json) and surfaces in classification.
"""

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data" / "raw" / "aikosh"
OUT = ROOT / "apps" / "api" / "data" / "category_demand.json"

# Keyword → ONDC domain mapping (mirrors the API's keyword classifier, wider net)
DOMAIN_KEYWORDS: dict[str, list[str]] = {
    "RET10": ["rice", "dal", "flour", "atta", "spice", "masala", "turmeric", "chilli",
              "oil", "ghee", "sugar", "tea", "coffee", "pickle", "papad", "snack",
              "honey", "jaggery", "millet", "grocery", "pulses", "peanut", "chikki",
              "sweet", "laddu", "murukku", "namkeen", "biscuit", "juice", "powder"],
    "RET11": ["food", "meal", "tiffin", "curry", "biryani", "cake", "bakery"],
    "RET12": ["saree", "sari", "kurta", "kurti", "dress", "cloth", "shirt", "fabric",
              "textile", "dupatta", "lehenga", "bag", "handbag", "jute", "cotton",
              "silk", "wool", "garment", "apparel", "chudidar", "blouse", "frock"],
    "RET13": ["soap", "shampoo", "cream", "lotion", "beauty", "cosmetic", "kajal",
              "essentials combo", "face", "hair oil", "herbal oil", "perfume"],
    "RET14": ["phone", "mobile", "charger", "led", "bulb", "battery", "electronic",
              "cable", "speaker", "earphone"],
    "RET15": ["mixer", "grinder", "fan", "cooker", "stove", "appliance", "iron"],
    "RET16": ["frame", "decor", "pot", "candle", "diya", "toy", "handicraft", "craft",
              "wooden", "bamboo", "terracotta", "pottery", "mat", "basket", "utensil",
              "steel", "brass", "furniture", "pillow", "bedsheet", "doormat", "photo",
              "cleaner", "phenyl", "detergent", "broom", "mop", "kitchen"],
    "RET18": ["ayurved", "herbal", "wellness", "yoga", "supplement", "organic",
              "medicinal", "immunity", "chyawanprash"],
}


def map_domain(*texts: str) -> str | None:
    blob = " ".join(t.lower() for t in texts if t)
    best, hits_best = None, 0
    for domain, kws in DOMAIN_KEYWORDS.items():
        hits = sum(1 for kw in kws if kw in blob)
        if hits > hits_best:
            best, hits_best = domain, hits
    return best


def main():
    orders_by_domain: dict[str, int] = defaultdict(int)
    value_by_domain: dict[str, float] = defaultdict(float)
    mapped = unmapped = 0

    # ── MyStore: product catalogue gives category text; Orders give volume ──
    wb = openpyxl.load_workbook(RAW / "ONDC-MyStore_App_Products_Orders.xlsx", read_only=True)
    prod_cat: dict[str, str] = {}
    for row in wb["Products"].iter_rows(min_row=2, values_only=True):
        name, cat = (str(row[1] or "").strip(), str(row[2] or "").strip())
        if name:
            prod_cat[name.lower()] = cat
    for row in wb["Orders"].iter_rows(min_row=2, values_only=True):
        name = str(row[1] or "").strip()
        try:
            value = float(row[2] or 0)
            count = int(float(row[3] or 0))
        except (ValueError, TypeError):
            continue
        domain = map_domain(name, prod_cat.get(name.lower(), ""))
        if domain:
            orders_by_domain[domain] += count
            value_by_domain[domain] += value
            mapped += count
        else:
            unmapped += count
    wb.close()

    # ── WowGeni: 137K raw order lines ──
    wb = openpyxl.load_workbook(RAW / "ONDC-WowGeni_App_Orders.xlsx", read_only=True)
    ws = wb["Sheet1"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[4] or "").strip()
        if not name:
            continue
        try:
            value = float(row[3] or 0)
        except (ValueError, TypeError):
            value = 0.0
        domain = map_domain(name)
        if domain:
            orders_by_domain[domain] += 1
            value_by_domain[domain] += value
            mapped += 1
        else:
            unmapped += 1
    wb.close()

    total = sum(orders_by_domain.values())
    ranked = sorted(orders_by_domain.items(), key=lambda x: -x[1])
    n = len(ranked)
    index = {}
    for i, (domain, count) in enumerate(ranked):
        level = "high" if i < max(1, n // 3) else ("medium" if i < max(2, 2 * n // 3) else "low")
        index[domain] = {
            "orders": count,
            "value_inr": round(value_by_domain[domain], 2),
            "share": round(count / total, 4) if total else 0,
            "level": level,
        }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({
        "meta": {
            "sources": [
                "AIKosh: ONDC-MyStore_App_Products_Orders.xlsx",
                "AIKosh: ONDC-WowGeni_App_Orders.xlsx",
            ],
            "total_orders_mapped": mapped,
            "orders_unmapped": unmapped,
            "method": "keyword mapping of real order lines to ONDC retail domains",
        },
        "domains": index,
    }, indent=2))
    print(f"mapped={mapped} unmapped={unmapped}")
    for d, v in index.items():
        print(f"  {d}: {v['orders']} orders, ₹{v['value_inr']:,.0f}, {v['level']}")
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
