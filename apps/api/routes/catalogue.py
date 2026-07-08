"""Catalogue Studio — ONDC-aligned product catalogue templates & upload.

The real-world SNP onboarding flow uses bulk-upload spreadsheets; ONDC itself
represents a seller's catalogue as a Beckn `on_search` catalog payload. This
module closes that loop:

  1. GET  /catalogue/template/{domain}  → multi-sheet XLSX template whose
     product columns follow the ONDC retail attribute spec for that domain
  2. POST /catalogue/upload             → parse + validate the filled sheet,
     auto-categorise each product, enrich the MSE profile for JodakAI,
     and generate the ONDC Beckn catalog JSON preview
"""

import io
import json
import re
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from database import MSE, AuditLog, User, get_db
from services.auth import get_current_user

router = APIRouter()

# ── ONDC retail attribute columns per domain (per ONDC RET enum spec) ──

BASE_COLUMNS = [
    ("product_name", "Product name as it should appear to buyers", True),
    ("description", "Short buyer-facing description", True),
    ("price_inr", "Selling price in ₹ (number only)", True),
    ("mrp_inr", "MRP in ₹", False),
    ("stock_qty", "Available quantity (number)", True),
    ("uom", "Unit of measure — see Reference sheet", False),
    ("hsn_code", "HSN code (4-8 digits)", False),
    ("sku_id", "Your internal product code", False),
    ("image_url", "Link to a product photo", False),
]

DOMAIN_COLUMNS: dict[str, list[tuple[str, str, bool]]] = {
    "RET10": [("net_weight", "e.g. 500 g / 1 kg", True),
              ("veg_nonveg", "veg / non-veg / egg", True),
              ("shelf_life_days", "Days before expiry", False),
              ("fssai_license_no", "14-digit FSSAI number", True)],
    "RET11": [("veg_nonveg", "veg / non-veg / egg", True),
              ("serves", "Serving size", False),
              ("fssai_license_no", "14-digit FSSAI number", True)],
    "RET12": [("colour", "See Reference sheet", True),
              ("size", "e.g. S / M / L / XL / Free Size", True),
              ("fabric", "See Reference sheet", True),
              ("gender", "male / female / unisex / boy / girl", True),
              ("pattern", "solid / printed / embroidered ...", False)],
    "RET13": [("net_quantity", "e.g. 100 ml", True),
              ("skin_type", "all / oily / dry / sensitive", False),
              ("shelf_life_months", "Months before expiry", False)],
    "RET14": [("brand", "Brand name", True),
              ("model", "Model number", False),
              ("warranty_months", "Warranty period in months", False),
              ("bis_crs_no", "BIS/CRS registration if notified item", False)],
    "RET15": [("brand", "Brand name", True),
              ("model", "Model number", False),
              ("energy_rating", "BEE star rating 1-5", False),
              ("warranty_months", "Warranty period in months", False)],
    "RET16": [("material", "e.g. brass / wood / bamboo / terracotta", True),
              ("colour", "See Reference sheet", False),
              ("dimensions_cm", "L x W x H in cm", False),
              ("handcrafted", "yes / no", False)],
    "RET18": [("net_quantity", "e.g. 60 capsules / 200 g", True),
              ("license_no", "FSSAI or AYUSH license number", True),
              ("shelf_life_months", "Months before expiry", False),
              ("dosage_form", "capsule / powder / oil / tea ...", False)],
}

REFERENCE_VALUES = {
    "uom": ["unit", "piece", "set", "pack", "kg", "g", "l", "ml", "dozen", "metre"],
    "colour": ["red", "blue", "green", "yellow", "black", "white", "pink", "orange",
               "purple", "brown", "beige", "maroon", "gold", "silver", "multicolour"],
    "fabric": ["cotton", "silk", "linen", "wool", "polyester", "rayon", "chiffon",
               "georgette", "khadi", "jute", "banarasi silk", "chanderi", "kanjeevaram"],
    "size": ["free size", "XS", "S", "M", "L", "XL", "XXL", "3XL"],
    "veg_nonveg": ["veg", "non-veg", "egg"],
}

_DOMAIN_LABELS = {
    "RET10": "Grocery", "RET11": "Food & Beverage", "RET12": "Fashion",
    "RET13": "Beauty & Personal Care", "RET14": "Electronics",
    "RET15": "Appliances", "RET16": "Home & Kitchen", "RET18": "Health & Wellness",
}


@router.get("/template/{domain}")
def download_template(
    domain: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Multi-sheet ONDC catalogue template for the given retail domain."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    domain = domain.upper()
    if domain not in DOMAIN_COLUMNS:
        raise HTTPException(status_code=404, detail="Unknown ONDC retail domain")

    columns = BASE_COLUMNS + DOMAIN_COLUMNS[domain]
    label = _DOMAIN_LABELS.get(domain, domain)

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1B4FCC")
    req_fill = PatternFill("solid", fgColor="E8680C")

    # Sheet 1 — Instructions
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = f"MSMEMate — ONDC {label} Catalogue Template ({domain})"
    ws["A1"].font = Font(bold=True, size=14)
    for i, line in enumerate([
        "1. Fill one product per row in the 'Products' sheet.",
        "2. Orange columns are MANDATORY for ONDC listing; blue are recommended.",
        "3. Use only the values listed in the 'Reference' sheet where indicated.",
        "4. Save the file and upload it back on the Catalogue page.",
        "5. MSMEMate validates every row, categorises your products, and",
        "   generates the ONDC (Beckn) catalog payload for your seller platform.",
        "", "इस शीट में हर पंक्ति में एक उत्पाद भरें। नारंगी कॉलम अनिवार्य हैं।",
    ], start=3):
        ws[f"A{i}"] = line

    # Sheet 2 — Products
    ws = wb.create_sheet("Products")
    for col, (name, hint, required) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = head_font
        c.fill = req_fill if required else head_fill
        ws.cell(row=2, column=col, value=hint).font = Font(italic=True, size=9, color="666666")
        ws.column_dimensions[c.column_letter].width = max(14, len(name) + 4)

    # Sheet 3 — Reference values
    ws = wb.create_sheet("Reference")
    for col, (field, values) in enumerate(REFERENCE_VALUES.items(), start=1):
        c = ws.cell(row=1, column=col, value=field)
        c.font = head_font
        c.fill = head_fill
        for row, v in enumerate(values, start=2):
            ws.cell(row=row, column=col, value=v)
        ws.column_dimensions[c.column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="MSMEMate_{domain}_catalogue_template.xlsx"'},
    )


# ── Upload, validate, categorise, and emit Beckn catalog ──────────────

class CatalogueItem(BaseModel):
    row: int
    product_name: str
    price_inr: float | None = None
    category_domain: str | None = None
    issues: list[str] = []


class CatalogueUploadResponse(BaseModel):
    total_rows: int
    valid_rows: int
    items: list[CatalogueItem]
    errors: list[str]
    beckn_catalog: dict
    profile_enriched: bool


def _rows_from_upload(file_bytes: bytes, filename: str) -> tuple[list[dict], list[str]]:
    """Parse Products rows from xlsx/csv into dicts keyed by header."""
    rows: list[dict] = []
    errors: list[str] = []
    if filename.lower().endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb["Products"] if "Products" in wb.sheetnames else wb[wb.sheetnames[0]]
        headers: list[str] = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_idx == 1:
                headers = [str(c or "").strip() for c in row]
                continue
            if r_idx == 2 and row and any("mandatory" in str(c or "").lower() or "e.g." in str(c or "") for c in row):
                continue  # hint row from our template
            values = {h: row[i] if i < len(row) else None for i, h in enumerate(headers) if h}
            if any(v not in (None, "") for v in values.values()):
                values["_row"] = r_idx
                rows.append(values)
        wb.close()
    else:
        import csv as _csv
        text = file_bytes.decode("utf-8", errors="replace")
        for r_idx, rec in enumerate(_csv.DictReader(io.StringIO(text)), start=2):
            rec = {k.strip(): v for k, v in rec.items() if k}
            if any((v or "").strip() for v in rec.values()):
                rec["_row"] = r_idx
                rows.append(rec)
    if not rows:
        errors.append("No product rows found — fill the Products sheet and re-upload.")
    return rows, errors


@router.post("/upload", response_model=CatalogueUploadResponse)
async def upload_catalogue(
    file: UploadFile = File(...),
    mse_id: int = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Validate a filled catalogue, categorise products, enrich the MSE
    profile for matching, and generate the ONDC Beckn catalog payload."""
    from services.classifier import _classify_with_keywords

    mse = db.query(MSE).get(mse_id)
    if not mse:
        raise HTTPException(status_code=404, detail="Business not found")

    file_bytes = await file.read()
    rows, errors = _rows_from_upload(file_bytes, file.filename or "catalogue.xlsx")

    items: list[CatalogueItem] = []
    beckn_items: list[dict] = []
    names: list[str] = []

    for rec in rows[:500]:
        row_no = int(rec.get("_row", 0))
        name = str(rec.get("product_name") or rec.get("Product name") or "").strip()
        issues: list[str] = []
        if not name:
            issues.append("product_name missing")
        price = None
        raw_price = rec.get("price_inr") or rec.get("price")
        if raw_price in (None, ""):
            issues.append("price_inr missing")
        else:
            try:
                price = float(re.sub(r"[^\d.]", "", str(raw_price)))
            except ValueError:
                issues.append("price_inr is not a number")
        qty = rec.get("stock_qty")
        if qty in (None, ""):
            issues.append("stock_qty missing")

        domain = None
        if name:
            preds = _classify_with_keywords(f"{name} {rec.get('description') or ''}")
            if preds and preds[0]["confidence"] > 0:
                domain = preds[0]["domain"]
            names.append(name)

        items.append(CatalogueItem(
            row=row_no, product_name=name or f"(row {row_no})",
            price_inr=price, category_domain=domain, issues=issues,
        ))

        if name and not issues:
            beckn_items.append({
                "id": str(rec.get("sku_id") or f"item-{row_no}"),
                "descriptor": {"name": name,
                               "short_desc": str(rec.get("description") or "")[:120]},
                "price": {"currency": "INR", "value": f"{price:.2f}"},
                "category_id": domain or "RET10",
                "quantity": {"available": {"count": str(qty)}},
                "@ondc/org/returnable": True,
                "@ondc/org/available_on_cod": True,
            })

    valid = sum(1 for it in items if not it.issues)

    # ONDC Beckn catalog payload (on_search shape, retail 1.2)
    beckn_catalog = {
        "context": {
            "domain": "ONDC:RET",
            "action": "on_search",
            "country": "IND",
            "city": f"std:{(mse.pin_code or '000000')[:3]}",
            "timestamp": datetime.now(timezone.utc).isoformat(),
        },
        "message": {
            "catalog": {
                "bpp/descriptor": {"name": "(your seller platform fills this)"},
                "bpp/providers": [{
                    "id": mse.udyam_number,
                    "descriptor": {"name": mse.name},
                    "locations": [{
                        "city": mse.district or "",
                        "state": mse.state or "",
                        "area_code": mse.pin_code or "",
                    }],
                    "items": beckn_items,
                }],
            }
        },
    }

    # Enrich the profile so JodakAI matches on the real catalogue
    enriched = False
    if names:
        existing = {p.strip().lower() for p in (mse.products or "").split(",") if p.strip()}
        new = [n for n in names[:25] if n.lower() not in existing]
        if new:
            mse.products = ", ".join(
                [p for p in [(mse.products or "").strip().strip(",")] if p] + new
            )[:2000]
            enriched = True

    db.add(AuditLog(
        action="catalogue_uploaded",
        entity_type="mse",
        entity_id=mse.id,
        details=f"{len(rows)} rows, {valid} valid, {len(beckn_items)} Beckn items",
        performed_by=user.username,
    ))
    db.commit()

    return CatalogueUploadResponse(
        total_rows=len(rows),
        valid_rows=valid,
        items=items[:100],
        errors=errors,
        beckn_catalog=beckn_catalog,
        profile_enriched=enriched,
    )
